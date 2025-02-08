# import_words.py
import os
from openpyxl import load_workbook
from sqlalchemy import create_engine, Column, Integer, String, Text
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.orm import sessionmaker

# Строка подключения к PostgreSQL
DATABASE_URL = "postgresql://postgres:VZGrUiwRbVXVcsZnZJtUqCDBkVmVBFYl@autorack.proxy.rlwy.net:29113/railway"

# Создаем базовый класс модели
Base = declarative_base()

class Word(Base):
    __tablename__ = 'words'
    id = Column(Integer, primary_key=True)
    word_et = Column(String, nullable=False)          # Эстонское слово
    part_of_speech = Column(String, nullable=False)     # Часть речи
    translation = Column(String, nullable=False)        # Перевод
    ai_generated_text = Column(Text)                    # Пример использования (опционально)
    # Новые столбцы, которых ожидает бот:
    correct_answers = Column(Integer, default=0)
    incorrect_answers = Column(Integer, default=0)

# Создаем движок и сессию
engine = create_engine(DATABASE_URL)
SessionLocal = sessionmaker(bind=engine)

def init_db():
    """Создает таблицы в БД, если их еще нет."""
    Base.metadata.create_all(engine)

def import_words_from_excel(filename):
    """
    Импортирует данные из Excel-файла в таблицу words.
    Предполагается, что первая строка – заголовки, данные начинаются со второй строки.
    """
    wb = load_workbook(filename)
    sheet = wb.active
    session = SessionLocal()

    for row in sheet.iter_rows(min_row=2, values_only=True):
        word_et = row[0]
        part_of_speech = row[1]
        translation = row[2]
        ai_generated_text = row[3] if len(row) > 3 else None

        word = Word(
            word_et=word_et,
            part_of_speech=part_of_speech,
            translation=translation,
            ai_generated_text=ai_generated_text
        )
        session.add(word)
    session.commit()
    session.close()
    print("Импорт завершён.")

if __name__ == '__main__':
    init_db()  # Создаем таблицы, если их нет
    import_words_from_excel('words.xlsx')
