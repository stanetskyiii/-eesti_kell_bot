import os
from openpyxl import load_workbook
from bot.database import SessionLocal, Word, init_db

def import_words_from_excel(excel_path):
    session = SessionLocal()
    wb = load_workbook(excel_path)
    sheet = wb.active

    # Предполагаем, что первая строка – заголовки:
    # Колонки: 1 - слово, 2 - часть речи, 3 - перевод, 4 - развёрнутый пример
    count = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        word_et, part_of_speech, translation, ai_generated_text = row[:4]

        # Проверяем, есть ли уже такое слово с данной частью речи
        existing = session.query(Word).filter_by(
            word_et=word_et,
            part_of_speech=part_of_speech
        ).first()

        if existing:
            print(f"Слово '{word_et}' с частью речи '{part_of_speech}' уже есть. Пропускаем.")
            continue

        new_word = Word(
            word_et=word_et,
            part_of_speech=part_of_speech,
            translation=translation,
            ai_generated_text=ai_generated_text
        )
        session.add(new_word)
        count += 1

    session.commit()
    session.close()
    print(f"Импортировано {count} новых слов.")

if __name__ == '__main__':
    # Инициализируем базу (создаёт таблицы, если их нет)
    init_db()
    # Путь к файлу Excel, который у вас находится по указанному пути
    excel_file = os.path.join("C:", os.sep, "Users", "Андрей", "Desktop", "телеграм эстонский", "words-final.xlsx")
    import_words_from_excel(excel_file)
